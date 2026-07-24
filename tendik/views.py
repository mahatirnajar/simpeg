from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
import io
from datetime import date
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import (
    Tendik, UnitKerja,
    RiwayatKepangkatanTendik, RiwayatJabatanFungsionalTendik,
    JabatanStrukturalTendik, RiwayatPendidikanTendik,
    TugasTambahanTendik, RiwayatBerhentiTendik,
    RiwayatStatusTendik, KeluargaTendik,
)
from .forms import (
    TendikForm, UnitKerjaForm,
    RiwayatKepangkatanTendikForm, RiwayatJabatanFungsionalTendikForm,
    JabatanStrukturalTendikForm, RiwayatPendidikanTendikForm,
    TugasTambahanTendikForm, RiwayatBerhentiTendikForm,
    RiwayatStatusTendikForm, KeluargaTendikForm,
)


def is_admin(user):
    return user.is_staff or user.is_superuser


# ─── DASHBOARD TENDIK ─────────────────────────────────────────────────────────

@login_required
def dashboard_tendik(request):
    search      = request.GET.get('q', '')
    unit_id     = request.GET.get('unit', '')
    status_f    = request.GET.get('status', '')
    jabatan_f   = request.GET.get('jabatan', '')

    qs = Tendik.objects.select_related('unit_kerja').prefetch_related(
        'riwayat_kepangkatan',
        'riwayat_jabatan_fungsional',
        'jabatan_struktural',
        'tugas_tambahan',
    )

    if search:
        qs = qs.filter(
            Q(nama_terang__icontains=search) |  
            Q(nip__icontains=search)            |
            Q(gelar_belakang__icontains=search) |
            Q(gelar_depan__icontains=search)
        )
    if unit_id:
        qs = qs.filter(unit_kerja_id=unit_id)
    if status_f:
        qs = qs.filter(status=status_f)

    rows = []
    no = 1
    for t in qs:
        # Keluarkan dari Dashboard jika sudah punya riwayat berhenti resmi,
        # ATAU sudah mencapai TMT Pensiun (BUP) — meskipun SK belum diproses,
        # tendik tetap otomatis pindah ke halaman Tendik Berhenti (status "Belum Diproses")
        if hasattr(t, 'riwayat_berhenti'):
            continue
        if t.sudah_pensiun:
            continue

        kp = t.kepangkatan_terakhir
        jf = t.jabatan_fungsional_terakhir
        js = t.jabatan_struktural_aktif
        tt = t.tugas_tambahan_aktif

        if jabatan_f:
            jf_match = jf and jabatan_f.lower() in (jf.nama_jabatan + ' ' + jf.jenjang).lower()
            js_match = js and jabatan_f.lower() in js.jabatan.lower()
            if not (jf_match or js_match):
                continue

        mk_k = t.masa_kerja_keseluruhan
        mk_g = t.masa_kerja_golongan
        mk_j = t.masa_kerja_jabatan
        mk_p = t.masa_kerja_pensiun

        rows.append({
            'no': no,
            'tendik': t,
            'kepangkatan': kp,
            'jabatan_fungsional': jf,
            'jabatan_struktural': js,
            'tugas_tambahan': tt,
            'mk_keseluruhan': mk_k,
            'mk_golongan':    mk_g,
            'mk_jabatan':     mk_j,
            'mk_pensiun':     mk_p,
            'sisa_mk_str':    t.sisa_masa_kerja_str,
        })
        no += 1

    per_page = int(request.GET.get('per_page', 10))
    if per_page > 9000:
        per_page = max(len(rows), 1)
    paginator = Paginator(rows, per_page or 10)
    page_obj  = paginator.get_page(request.GET.get('page'))

    # Stats — hanya tendik aktif (belum berhenti, belum lewat BUP)
    tendik_berhenti_ids = set(
        Tendik.objects.exclude(riwayat_berhenti__isnull=True).values_list('pk', flat=True)
    )
    today = date.today()
    tendik_sudah_bup_ids = set(
        Tendik.objects.filter(
            tmt_pensiun__isnull=False, tmt_pensiun__lte=today
        ).values_list('pk', flat=True)
    )
    exclude_ids = tendik_berhenti_ids | tendik_sudah_bup_ids

    total       = Tendik.objects.exclude(pk__in=exclude_ids).count()
    pns_count   = Tendik.objects.exclude(pk__in=exclude_ids).filter(status='PNS').count()
    cpns_count  = Tendik.objects.exclude(pk__in=exclude_ids).filter(status='CPNS').count()
    pppk_count  = Tendik.objects.exclude(pk__in=exclude_ids).filter(status='PPPK').count()
    by_unit     = UnitKerja.objects.annotate(jumlah=Count('tendik')).filter(jumlah__gt=0).order_by('-jumlah')

    context = {
        'page_obj':      page_obj,
        'per_page':      str(per_page),
        'total_tendik':  total,
        'pns_count':     pns_count,
        'cpns_count':    cpns_count,
        'pppk_count':    pppk_count,
        'by_unit':       by_unit,
        'unit_list':     UnitKerja.objects.all().order_by('nama'),
        'search':        search,
        'unit_id':       unit_id,
        'status_filter': status_f,
        'jabatan_filter': jabatan_f,
        'is_admin':      is_admin(request.user),
    }
    return render(request, 'tendik/dashboard_tendik.html', context)


# ─── NAIK PANGKAT TENDIK ──────────────────────────────────────────────────────

@login_required
def naik_pangkat_tendik_list(request):
    search  = request.GET.get('q', '')
    unit_id = request.GET.get('unit', '')

    today = date.today()
    try:
        batas_tmt = today.replace(year=today.year - 4)
    except ValueError:
        batas_tmt = today.replace(month=2, day=28, year=today.year - 4)

    qs = Tendik.objects.select_related('unit_kerja').prefetch_related(
        'riwayat_kepangkatan', 'riwayat_jabatan_fungsional'
    )

    if search:
        qs = qs.filter(
            Q(nama_terang__icontains=search) |
            Q(nip__icontains=search) |
            Q(gelar_belakang__icontains=search) |
            Q(gelar_depan__icontains=search)
        )
    if unit_id:
        qs = qs.filter(unit_kerja_id=unit_id)

    def lama_menjabat(tmt, hari_ini):
        years = hari_ini.year - tmt.year
        months = hari_ini.month - tmt.month
        if hari_ini.day < tmt.day:
            months -= 1
        if months < 0:
            years -= 1
            months += 12
        return years, months

    rows = []
    no = 1
    for t in qs:
        if hasattr(t, 'riwayat_berhenti'):
            continue

        kp = t.kepangkatan_terakhir
        if not kp or not kp.tmt:
            continue
        if kp.tmt > batas_tmt:
            continue

        lama_tahun, lama_bulan = lama_menjabat(kp.tmt, today)

        rows.append({
            'no': no,
            'tendik': t,
            'kepangkatan': kp,
            'jabatan_fungsional': t.jabatan_fungsional_terakhir,
            'lama_tahun': lama_tahun,
            'lama_bulan': lama_bulan,
        })
        no += 1

    per_page = int(request.GET.get('per_page', 25))
    if per_page > 9000:
        per_page = len(rows) or 1
    paginator = Paginator(rows, per_page or 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'total_eligible': len(rows),
        'unit_list': UnitKerja.objects.all().order_by('nama'),
        'search': search,
        'unit_id': unit_id,
        'is_admin': is_admin(request.user),
    }
    return render(request, 'tendik/naik_pangkat_list_tendik.html', context)


# ─── TENDIK BERHENTI / PENSIUN ─────────────────────────────────────────────────

@login_required
def tendik_berhenti_list(request):
    search  = request.GET.get('q', '')
    unit_id = request.GET.get('unit', '')

    today = date.today()

    qs = Tendik.objects.select_related('unit_kerja').prefetch_related(
        'riwayat_jabatan_fungsional'
    )

    if search:
        qs = qs.filter(
            Q(nama_terang__icontains=search) |
            Q(nip__icontains=search) |
            Q(gelar_belakang__icontains=search) |
            Q(gelar_depan__icontains=search)
        )
    if unit_id:
        qs = qs.filter(unit_kerja_id=unit_id)

    rows = []
    no = 1
    count_belum_diproses = 0
    count_sudah_diproses = 0

    for t in qs:
        if not t.sudah_pensiun:
            continue

        riwayat_berhenti = getattr(t, 'riwayat_berhenti', None)
        sudah_diproses = riwayat_berhenti is not None

        if sudah_diproses:
            count_sudah_diproses += 1
        else:
            count_belum_diproses += 1

        selisih_bulan = (today.year - t.tmt_pensiun.year) * 12 + (today.month - t.tmt_pensiun.month)
        if today.day < t.tmt_pensiun.day:
            selisih_bulan -= 1
        lewat_tahun, lewat_bulan = divmod(max(selisih_bulan, 0), 12)

        rows.append({
            'no': no,
            'tendik': t,
            'tmt_pensiun': t.tmt_pensiun,
            'lewat_tahun': lewat_tahun,
            'lewat_bulan': lewat_bulan,
            'jabatan_fungsional': t.jabatan_fungsional_terakhir,
            'riwayat_berhenti': riwayat_berhenti,
            'sudah_diproses': sudah_diproses,
        })
        no += 1

    rows.sort(key=lambda r: (r['sudah_diproses'], -(r['lewat_tahun'] * 12 + r['lewat_bulan'])))
    for i, r in enumerate(rows, start=1):
        r['no'] = i

    per_page = int(request.GET.get('per_page', 25))
    if per_page > 9000:
        per_page = len(rows) or 1
    paginator = Paginator(rows, per_page or 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'total_berhenti': len(rows),
        'count_belum_diproses': count_belum_diproses,
        'count_sudah_diproses': count_sudah_diproses,
        'unit_list': UnitKerja.objects.all().order_by('nama'),
        'search': search,
        'unit_id': unit_id,
        'is_admin': is_admin(request.user),
    }
    return render(request, 'tendik/tendik_berhenti_list.html', context)


# ─── PROSES BERHENTI/PENSIUN ───────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def tendik_berhenti_add(request, tendik_pk):
    tendik = get_object_or_404(Tendik, pk=tendik_pk)
    if hasattr(tendik, 'riwayat_berhenti'):
        messages.warning(request, 'Tendik ini sudah tercatat berhenti sebelumnya.')
        return redirect('tendik_detail', pk=tendik_pk)

    if request.method == 'POST':
        form = RiwayatBerhentiTendikForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tendik = tendik
            obj.save()
            messages.success(request, 'Data pemberhentian/pensiun berhasil disimpan.')
            return redirect('tendik_detail', pk=tendik_pk)
    else:
        form = RiwayatBerhentiTendikForm(initial={
            'alasan': 'PENSIUN',
            'tanggal': tendik.tmt_pensiun,
        })
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': tendik, 'title': 'Proses Pemberhentian/Pensiun'
    })


# ─── RIWAYAT STATUS TENDIK ─────────────────────────────────────────────────────

@login_required
def status_tendik_add(request, tendik_id):
    tendik = get_object_or_404(Tendik, pk=tendik_id)
    if request.method == "POST":
        form = RiwayatStatusTendikForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tendik = tendik
            obj.save()
            messages.success(request, "Riwayat status berhasil ditambahkan.")
            return redirect('tendik_detail', pk=tendik.pk)
    else:
        form = RiwayatStatusTendikForm()
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': tendik, 'title': 'Tambah Riwayat Status'
    })


@login_required
def status_tendik_edit(request, pk):
    obj = get_object_or_404(RiwayatStatusTendik, pk=pk)
    if request.method == "POST":
        form = RiwayatStatusTendikForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Riwayat status berhasil diperbarui.")
            return redirect('tendik_detail', pk=obj.tendik_id)
    else:
        form = RiwayatStatusTendikForm(instance=obj)
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': obj.tendik, 'title': 'Edit Riwayat Status'
    })


@login_required
def status_tendik_delete(request, pk):
    obj = get_object_or_404(RiwayatStatusTendik, pk=pk)
    tendik_pk = obj.tendik_id
    if request.method == 'POST':
        obj.delete()
        messages.success(request, "Riwayat status berhasil dihapus.")
        return redirect('tendik_detail', pk=tendik_pk)
    return render(request, 'tendik/confirm_delete_tendik.html', {
        'obj': obj, 'title': 'Hapus Riwayat Status'
    })


# ─── KELUARGA TENDIK ────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def keluarga_tendik_add(request, tendik_id):
    tendik = get_object_or_404(Tendik, pk=tendik_id)
    if request.method == "POST":
        form = KeluargaTendikForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tendik = tendik
            obj.save()
            messages.success(request, "Data keluarga berhasil ditambahkan.")
            return redirect('tendik_detail', pk=tendik.pk)
    else:
        form = KeluargaTendikForm()
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': tendik, 'title': 'Tambah Data Keluarga'
    })


@login_required
@user_passes_test(is_admin)
def keluarga_tendik_edit(request, pk):
    obj = get_object_or_404(KeluargaTendik, pk=pk)
    if request.method == "POST":
        form = KeluargaTendikForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Data keluarga berhasil diperbarui.")
            return redirect('tendik_detail', pk=obj.tendik_id)
    else:
        form = KeluargaTendikForm(instance=obj)
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': obj.tendik, 'title': 'Edit Data Keluarga'
    })


@login_required
@user_passes_test(is_admin)
def keluarga_tendik_delete(request, pk):
    obj = get_object_or_404(KeluargaTendik, pk=pk)
    tendik_pk = obj.tendik_id
    if request.method == 'POST':
        obj.delete()
        messages.success(request, "Data keluarga berhasil dihapus.")
        return redirect('tendik_detail', pk=tendik_pk)
    return render(request, 'tendik/confirm_delete_tendik.html', {
        'obj': obj, 'title': 'Hapus Data Keluarga'
    })


# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Tendik"

    # ── Styles ──────────────────────────────────────────────────────────────
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    sub_fill = PatternFill("solid", start_color="374151")
    sub_font = Font(name="Arial", bold=True, color="FFFFFF", size=8)
    alt_fill = PatternFill("solid", start_color="EBF3FB")
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="BFBFBF")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = border
        return c

    def scell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = sub_font; c.fill = sub_fill
        c.alignment = center; c.border = border
        return c

    # ── Baris 1: Judul ──────────────────────────────────────────────────────
    TOTAL_COLS = 34
    ws.merge_cells(f"A1:{get_column_letter(TOTAL_COLS)}1")
    tc = ws["A1"]
    tc.value     = "DATA TENAGA KEPENDIDIKAN UNIVERSITAS TADULAKO"
    tc.font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    tc.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Layout kolom ────────────────────────────────────────────────────────
    single_cols = {
        1:  ("NO",               6),
        2:  ("NIP",             20),
        3:  ("NAMA",            32),
        4:  ("UNIT KERJA",      20),
        5:  ("STATUS",          10),
        6:  ("L/P",              6),
        7:  ("PANGKAT",         20),
        8:  ("GOL.",             8),
        9:  ("TMT",             13),
        10: ("JAB. FUNGSIONAL", 28),
        11: ("JAB. STRUKTURAL", 24),
        12: ("ESELON",           9),
        23: ("BAGIAN",          20),
        24: ("KARPEG",          14),
        25: ("TMT CPNS",        13),
        26: ("TMT PNS",         13),
        27: ("AGAMA",           12),
        28: ("KEPAKARAN",       22),
        29: ("IJ. BKN",         10),
        30: ("IJ. BORANG",      11),
        31: ("TMP LAHIR",       16),
        32: ("TGL LAHIR",       13),
        33: ("USIA",             7),
        34: ("PENSIUN",          9),
    }
    grouped_cols = {
        13: "MK GOL",
        15: "MK JAB",
        17: "MK KESELURUHAN",
        19: "MK PENSIUN (SISA)",
    }
    single_cols[21] = ("TMT PENSIUN", 14)
    single_cols[22] = ("SISA MK",     14)

    for col, (label, width) in single_cols.items():
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        hcell(2, col, label)
        ws.column_dimensions[get_column_letter(col)].width = width

    for start_col, label in grouped_cols.items():
        ws.merge_cells(start_row=2, start_column=start_col,
                       end_row=2,   end_column=start_col + 1)
        hcell(2, start_col, label)
        scell(3, start_col,     "Thn")
        scell(3, start_col + 1, "Bln")
        ws.column_dimensions[get_column_letter(start_col)].width     = 7
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 7

    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"

    # ── Data rows ────────────────────────────────────────────────────────────
    qs = (
        Tendik.objects
        .select_related("unit_kerja")
        .prefetch_related(
            "riwayat_kepangkatan",
            "riwayat_jabatan_fungsional",
            "jabatan_struktural",
        )
        .order_by("nama_terang")
    )

    def fmt_date(d):
        return d.strftime("%d-%m-%Y") if d else "-"

    def vd(v):
        return v if v else "-"

    center_cols = {0, 5, 7, 8, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 32, 33}

    for row_num, t in enumerate(qs, start=1):
        row_idx = row_num + 3
        fill    = alt_fill if row_num % 2 == 0 else None

        kp  = t.kepangkatan_terakhir
        jf  = t.jabatan_fungsional_terakhir
        js  = t.jabatan_struktural_aktif

        mk_g = t.masa_kerja_golongan
        mk_j = t.masa_kerja_jabatan
        mk_k = t.masa_kerja_keseluruhan
        mk_p = t.masa_kerja_pensiun

        jabf_str = jf.nama_dan_jenjang if jf else "-"

        row_data = [
            row_num,
            vd(t.nip),
            t.nama_lengkap,
            t.nama_unit,
            t.get_status_display(),
            t.jenis_kelamin,
            kp.pangkat   if kp else "-",
            kp.golongan  if kp else "-",
            fmt_date(kp.tmt) if kp else "-",
            jabf_str,
            js.jabatan   if js else "-",
            vd(js.eselon) if js else "-",
            mk_g[0], mk_g[1],
            mk_j[0], mk_j[1],
            mk_k[0], mk_k[1],
            mk_p[0], mk_p[1],
            fmt_date(t.tmt_pensiun),
            t.sisa_masa_kerja_str,
            vd(t.bagian),
            vd(t.karpeg),
            fmt_date(t.tmt_cpns),
            fmt_date(t.tmt_pns),
            vd(t.agama),
            vd(t.bidang_keahlian),
            vd(t.tingkat_ijazah_bkn),
            vd(t.tingkat_ijazah_borang),
            vd(t.tempat_lahir),
            fmt_date(t.tanggal_lahir),
            t.usia_saat_ini or "-",
            t.tahun_pensiun or "-",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = center if (col_idx - 1) in center_cols else left
            cell.border    = border
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Data_Tendik_UNTAD.xlsx"'
    return response


# ─── DETAIL TENDIK ────────────────────────────────────────────────────────────

@login_required
def tendik_detail(request, pk):
    tendik = get_object_or_404(
        Tendik.objects.select_related('unit_kerja').prefetch_related(
            'riwayat_kepangkatan',
            'riwayat_jabatan_fungsional',
            'jabatan_struktural',
            'riwayat_pendidikan',
            'tugas_tambahan',
            'riwayat_status',
            'keluarga',
        ),
        pk=pk
    )
    return render(request, 'tendik/tendik_detail.html', {
        'tendik': tendik,
        'is_admin': is_admin(request.user),
        'mk_keseluruhan': tendik.masa_kerja_keseluruhan,
        'mk_golongan':    tendik.masa_kerja_golongan,
        'mk_jabatan':     tendik.masa_kerja_jabatan,
        'mk_pensiun':     tendik.masa_kerja_pensiun,
    })


# ─── TENDIK CRUD ──────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def tendik_create(request):
    if request.method == 'POST':
        form = TendikForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, f'Data tendik {obj.nama_lengkap} berhasil disimpan.')
            return redirect('tendik_detail', pk=obj.pk)
    else:
        form = TendikForm()
    return render(request, 'tendik/tendik_form.html', {
        'form': form, 'title': 'Tambah Tendik Baru'
    })


@login_required
@user_passes_test(is_admin)
def tendik_edit(request, pk):
    tendik = get_object_or_404(Tendik, pk=pk)
    if request.method == 'POST':
        form = TendikForm(request.POST, instance=tendik)
        if form.is_valid():
            form.save()
            messages.success(request, f'Data {tendik.nama_lengkap} berhasil diperbarui.')
            return redirect('tendik_detail', pk=tendik.pk)
    else:
        form = TendikForm(instance=tendik)
    return render(request, 'tendik/tendik_form.html', {
        'form': form, 'tendik': tendik,
        'title': f'Edit: {tendik.nama_lengkap}'
    })


@login_required
@user_passes_test(is_admin)
def tendik_delete(request, pk):
    tendik = get_object_or_404(Tendik, pk=pk)
    if request.method == 'POST':
        nama = tendik.nama_lengkap
        tendik.delete()
        messages.success(request, f'Data tendik {nama} berhasil dihapus.')
        return redirect('dashboard_tendik')
    return render(request, 'tendik/confirm_delete_tendik.html', {
        'obj': tendik, 'title': 'Hapus Tendik'
    })


# ─── KEPANGKATAN ──────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def kepangkatan_tendik_add(request, tendik_pk):
    tendik = get_object_or_404(Tendik, pk=tendik_pk)
    if request.method == 'POST':
        form = RiwayatKepangkatanTendikForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tendik = tendik
            obj.save()
            messages.success(request, 'Riwayat kepangkatan ditambahkan.')
            return redirect('tendik_detail', pk=tendik_pk)
    else:
        form = RiwayatKepangkatanTendikForm()
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': tendik, 'title': 'Tambah Riwayat Kepangkatan'
    })


@login_required
@user_passes_test(is_admin)
def kepangkatan_tendik_edit(request, pk):
    obj = get_object_or_404(RiwayatKepangkatanTendik, pk=pk)
    if request.method == 'POST':
        form = RiwayatKepangkatanTendikForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Riwayat kepangkatan diperbarui.')
            return redirect('tendik_detail', pk=obj.tendik_id)
    else:
        form = RiwayatKepangkatanTendikForm(instance=obj)
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': obj.tendik, 'title': 'Edit Riwayat Kepangkatan'
    })


@login_required
@user_passes_test(is_admin)
def kepangkatan_tendik_delete(request, pk):
    obj = get_object_or_404(RiwayatKepangkatanTendik, pk=pk)
    tendik_pk = obj.tendik_id
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Dihapus.')
        return redirect('tendik_detail', pk=tendik_pk)
    return render(request, 'tendik/confirm_delete_tendik.html', {
        'obj': obj, 'title': 'Hapus Kepangkatan'
    })


# ─── JABATAN FUNGSIONAL ───────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def jabatan_tendik_add(request, tendik_pk):
    tendik = get_object_or_404(Tendik, pk=tendik_pk)
    if request.method == 'POST':
        form = RiwayatJabatanFungsionalTendikForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tendik = tendik
            obj.save()
            messages.success(request, 'Riwayat jabatan fungsional ditambahkan.')
            return redirect('tendik_detail', pk=tendik_pk)
    else:
        form = RiwayatJabatanFungsionalTendikForm()
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': tendik, 'title': 'Tambah Jabatan Fungsional'
    })


@login_required
@user_passes_test(is_admin)
def jabatan_tendik_edit(request, pk):
    obj = get_object_or_404(RiwayatJabatanFungsionalTendik, pk=pk)
    if request.method == 'POST':
        form = RiwayatJabatanFungsionalTendikForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Diperbarui.')
            return redirect('tendik_detail', pk=obj.tendik_id)
    else:
        form = RiwayatJabatanFungsionalTendikForm(instance=obj)
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': obj.tendik, 'title': 'Edit Jabatan Fungsional'
    })


@login_required
@user_passes_test(is_admin)
def jabatan_tendik_delete(request, pk):
    obj = get_object_or_404(RiwayatJabatanFungsionalTendik, pk=pk)
    tendik_pk = obj.tendik_id
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Dihapus.')
        return redirect('tendik_detail', pk=tendik_pk)
    return render(request, 'tendik/confirm_delete_tendik.html', {
        'obj': obj, 'title': 'Hapus Jabatan Fungsional'
    })


# ─── JABATAN STRUKTURAL ───────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def jabatan_struktural_add(request, tendik_pk):
    tendik = get_object_or_404(Tendik, pk=tendik_pk)
    if request.method == 'POST':
        form = JabatanStrukturalTendikForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tendik = tendik
            obj.save()
            messages.success(request, 'Jabatan struktural ditambahkan.')
            return redirect('tendik_detail', pk=tendik_pk)
    else:
        form = JabatanStrukturalTendikForm()
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': tendik, 'title': 'Tambah Jabatan Struktural'
    })


@login_required
@user_passes_test(is_admin)
def jabatan_struktural_edit(request, pk):
    obj = get_object_or_404(JabatanStrukturalTendik, pk=pk)
    if request.method == 'POST':
        form = JabatanStrukturalTendikForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Diperbarui.')
            return redirect('tendik_detail', pk=obj.tendik_id)
    else:
        form = JabatanStrukturalTendikForm(instance=obj)
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': obj.tendik, 'title': 'Edit Jabatan Struktural'
    })


@login_required
@user_passes_test(is_admin)
def jabatan_struktural_delete(request, pk):
    obj = get_object_or_404(JabatanStrukturalTendik, pk=pk)
    tendik_pk = obj.tendik_id
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Dihapus.')
        return redirect('tendik_detail', pk=tendik_pk)
    return render(request, 'tendik/confirm_delete_tendik.html', {
        'obj': obj, 'title': 'Hapus Jabatan Struktural'
    })


# ─── PENDIDIKAN ───────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def pendidikan_tendik_add(request, tendik_pk):
    tendik = get_object_or_404(Tendik, pk=tendik_pk)
    if request.method == 'POST':
        form = RiwayatPendidikanTendikForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tendik = tendik
            obj.save()
            messages.success(request, 'Riwayat pendidikan ditambahkan.')
            return redirect('tendik_detail', pk=tendik_pk)
    else:
        form = RiwayatPendidikanTendikForm()
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': tendik, 'title': 'Tambah Riwayat Pendidikan'
    })


@login_required
@user_passes_test(is_admin)
def pendidikan_tendik_delete(request, pk):
    obj = get_object_or_404(RiwayatPendidikanTendik, pk=pk)
    tendik_pk = obj.tendik_id
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Dihapus.')
        return redirect('tendik_detail', pk=tendik_pk)
    return render(request, 'tendik/confirm_delete_tendik.html', {
        'obj': obj, 'title': 'Hapus Pendidikan'
    })


# ─── TUGAS TAMBAHAN ───────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def tugas_tendik_add(request, tendik_pk):
    tendik = get_object_or_404(Tendik, pk=tendik_pk)
    if request.method == 'POST':
        form = TugasTambahanTendikForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tendik = tendik
            obj.save()
            messages.success(request, 'Tugas tambahan ditambahkan.')
            return redirect('tendik_detail', pk=tendik_pk)
    else:
        form = TugasTambahanTendikForm()
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': tendik, 'title': 'Tambah Tugas Tambahan'
    })


@login_required
@user_passes_test(is_admin)
def tugas_tendik_edit(request, pk):
    obj = get_object_or_404(TugasTambahanTendik, pk=pk)
    if request.method == 'POST':
        form = TugasTambahanTendikForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Diperbarui.')
            return redirect('tendik_detail', pk=obj.tendik_id)
    else:
        form = TugasTambahanTendikForm(instance=obj)
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'tendik': obj.tendik, 'title': 'Edit Tugas Tambahan'
    })


@login_required
@user_passes_test(is_admin)
def tugas_tendik_delete(request, pk):
    obj = get_object_or_404(TugasTambahanTendik, pk=pk)
    tendik_pk = obj.tendik_id
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Dihapus.')
        return redirect('tendik_detail', pk=tendik_pk)
    return render(request, 'tendik/confirm_delete_tendik.html', {
        'obj': obj, 'title': 'Hapus Tugas Tambahan'
    })


# ─── UNIT KERJA ──────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def unit_kerja_list(request):
    units = UnitKerja.objects.annotate(jumlah_tendik=Count('tendik')).order_by('nama')
    return render(request, 'tendik/unit_kerja_list.html', {'unit_list': units})


@login_required
@user_passes_test(is_admin)
def unit_kerja_create(request):
    if request.method == 'POST':
        form = UnitKerjaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unit kerja ditambahkan.')
            return redirect('unit_kerja_list')
    else:
        form = UnitKerjaForm()
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'title': 'Tambah Unit Kerja'
    })


@login_required
@user_passes_test(is_admin)
def unit_kerja_edit(request, pk):
    unit = get_object_or_404(UnitKerja, pk=pk)
    if request.method == 'POST':
        form = UnitKerjaForm(request.POST, instance=unit)
        if form.is_valid():
            form.save()
            messages.success(request, 'Unit kerja diperbarui.')
            return redirect('unit_kerja_list')
    else:
        form = UnitKerjaForm(instance=unit)
    return render(request, 'tendik/riwayat_form_tendik.html', {
        'form': form, 'title': f'Edit Unit Kerja: {unit.nama}'
    })