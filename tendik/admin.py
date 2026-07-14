"""
admin.py — Tendik UNTAD
Fitur:
  - Search, filter, list_display lengkap
  - Inline untuk semua relasi (kepangkatan, jabatan, pendidikan, cuti, dll)
  - Export Excel (xlsx) via action — tanpa library tambahan selain openpyxl
  - Kolom kalkulasi (@property) ditampilkan di list & readonly_fields
"""

import openpyxl
from datetime import date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html

from .models import (
    UnitKerja, ReferensiJabatan, Tendik,
    RiwayatKepangkatanTendik, RiwayatJabatanFungsionalTendik,
    JabatanStrukturalTendik, TugasTambahanTendik,
    RiwayatPendidikanTendik, DetailPPPK,
    CutiTendik, RiwayatBerhentiTendik,
)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def _style_header(cell):
    cell.font      = Font(bold=True, color='FFFFFF', size=10)
    cell.fill      = PatternFill('solid', fgColor='1F4E79')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

def _style_data(cell, row_idx):
    cell.fill      = PatternFill('solid', fgColor='DCE6F1' if row_idx % 2 == 0 else 'FFFFFF')
    cell.alignment = Alignment(vertical='center', wrap_text=False)
    cell.border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

def _autofit(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(max_len + 2, min_width), max_width)

def _export_response(wb, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────────────────────
# ACTION EXPORT — dipakai di TendikAdmin
# ─────────────────────────────────────────────────────────────────────────────

def export_tendik_excel(modeladmin, request, queryset):
    """Export data Tendik terpilih ke Excel (.xlsx)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Tendik'
    ws.freeze_panes = 'C3'

    headers = [
        'No', 'NIP', 'Nama Lengkap', 'Nama Terang', 'Gelar Depan', 'Gelar Belakang',
        'Jenis Kelamin', 'Agama', 'Tempat Lahir', 'Tanggal Lahir', 'Usia Saat Ini',
        'Status', 'Unit Kerja', 'Bagian',
        'Pangkat Terakhir', 'Golongan',
        'Jabatan Fungsional', 'Jenjang Jabatan',
        'TMT CPNS', 'TMT PNS',
        'MK Keseluruhan (Thn)', 'MK Keseluruhan (Bln)',
        'MK Golongan (Thn)', 'MK Golongan (Bln)',
        'MK Jabatan (Thn)', 'MK Jabatan (Bln)',
        'Usia Pensiun', 'Tahun Pensiun', 'TMT Pensiun', 'Sisa MK',
        'TMT KGB',
        'Ijazah Tertinggi', 'Bidang Keahlian',
        'No. HP', 'Email',
        'KARPEG', 'No. KTP',
    ]

    # Baris header
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        _style_header(cell)
    ws.row_dimensions[1].height = 30

    # Data
    for row_idx, t in enumerate(queryset.select_related(
        'unit_kerja',
    ).prefetch_related(
        'riwayat_kepangkatan', 'riwayat_jabatan_fungsional',
    ), start=2):

        kpk  = t.kepangkatan_terakhir
        jabf = t.jabatan_fungsional_terakhir
        mk_k = t.masa_kerja_keseluruhan
        mk_g = t.masa_kerja_golongan
        mk_j = t.masa_kerja_jabatan

        row = [
            row_idx - 1,
            t.nip or '',
            t.nama_lengkap,
            t.nama_terang,
            t.gelar_depan,
            t.gelar_belakang,
            t.get_jenis_kelamin_display(),
            t.agama,
            t.tempat_lahir,
            t.tanggal_lahir,
            t.usia_saat_ini,
            t.get_status_display(),
            t.nama_unit,
            t.bagian,
            kpk.pangkat  if kpk else '',
            kpk.golongan if kpk else '',
            jabf.nama_jabatan if jabf else '',
            jabf.jenjang      if jabf else '',
            t.tmt_cpns,
            t.tmt_pns,
            mk_k[0], mk_k[1],
            mk_g[0], mk_g[1],
            mk_j[0], mk_j[1],
            t.usia_pensiun,
            t.tahun_pensiun,
            t.tmt_pensiun,
            t.sisa_masa_kerja_str,
            t.tmt_kgb,
            t.tingkat_ijazah_bkn,
            t.bidang_keahlian,
            t.no_hp,
            t.email,
            t.karpeg or '',
            t.no_ktp or '',
        ]

        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _style_data(cell, row_idx)
            if isinstance(val, date):
                cell.number_format = 'DD/MM/YYYY'

    _autofit(ws)
    today = date.today().strftime('%Y%m%d')
    return _export_response(wb, f'Data_Tendik_{today}.xlsx')

export_tendik_excel.short_description = '📥 Export Excel (Tendik terpilih)'


def export_tendik_kepangkatan_excel(modeladmin, request, queryset):
    """Export riwayat kepangkatan Tendik terpilih."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kepangkatan'

    headers = ['No', 'NIP', 'Nama Lengkap', 'Status', 'Unit Kerja',
               'Pangkat', 'Golongan', 'TMT', 'Keterangan']
    for col, h in enumerate(headers, 1):
        _style_header(ws.cell(row=1, column=col, value=h))

    row_idx = 2
    for t in queryset.select_related('unit_kerja').prefetch_related('riwayat_kepangkatan'):
        for kpk in t.riwayat_kepangkatan.order_by('-tmt'):
            row = [row_idx - 1, t.nip or '', t.nama_lengkap, t.status,
                   t.nama_unit, kpk.pangkat, kpk.golongan, kpk.tmt, kpk.keterangan]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                _style_data(cell, row_idx)
                if isinstance(val, date):
                    cell.number_format = 'DD/MM/YYYY'
            row_idx += 1

    _autofit(ws)
    today = date.today().strftime('%Y%m%d')
    return _export_response(wb, f'Kepangkatan_Tendik_{today}.xlsx')

export_tendik_kepangkatan_excel.short_description = '📥 Export Excel (Riwayat Kepangkatan)'


def export_tendik_pensiun_excel(modeladmin, request, queryset):
    """Export data pensiun Tendik terpilih (diurutkan TMT Pensiun)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Pensiun'

    headers = ['No', 'NIP', 'Nama Lengkap', 'Status', 'Unit Kerja',
               'Tgl Lahir', 'Usia Saat Ini', 'Usia Pensiun',
               'Tahun Pensiun', 'TMT Pensiun', 'Sisa Masa Kerja']
    for col, h in enumerate(headers, 1):
        _style_header(ws.cell(row=1, column=col, value=h))

    # Urutkan berdasarkan TMT pensiun terdekat
    tendik_list = list(queryset.select_related('unit_kerja'))
    tendik_list.sort(key=lambda t: t.tmt_pensiun or date(9999, 12, 31))

    for row_idx, t in enumerate(tendik_list, 2):
        row = [
            row_idx - 1, t.nip or '', t.nama_lengkap, t.status, t.nama_unit,
            t.tanggal_lahir, t.usia_saat_ini, t.usia_pensiun,
            t.tahun_pensiun, t.tmt_pensiun, t.sisa_masa_kerja_str,
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _style_data(cell, row_idx)
            if isinstance(val, date):
                cell.number_format = 'DD/MM/YYYY'

    _autofit(ws)
    today = date.today().strftime('%Y%m%d')
    return _export_response(wb, f'Pensiun_Tendik_{today}.xlsx')

export_tendik_pensiun_excel.short_description = '📥 Export Excel (Data Pensiun)'


# ─────────────────────────────────────────────────────────────────────────────
# INLINE
# ─────────────────────────────────────────────────────────────────────────────

class KepangkatanInline(admin.TabularInline):
    model   = RiwayatKepangkatanTendik
    extra   = 1
    fields  = ('pangkat', 'golongan', 'tmt', 'keterangan')
    ordering = ('-tmt',)


class JabatanFungsionalInline(admin.TabularInline):
    model   = RiwayatJabatanFungsionalTendik
    extra   = 1
    fields  = ('jenis', 'nama_jabatan', 'jenjang', 'tmt', 'angka_kredit', 'keterangan')
    ordering = ('-tmt',)


class JabatanStrukturalInline(admin.TabularInline):
    model   = JabatanStrukturalTendik
    extra   = 0
    fields  = ('jabatan', 'eselon', 'unit_kerja', 'tmt_jabatan', 'is_aktif', 'no_sk')
    ordering = ('-tmt_jabatan',)


class TugasTambahanInline(admin.TabularInline):
    model   = TugasTambahanTendik
    extra   = 0
    fields  = ('jabatan', 'tmt_jabatan', 'is_aktif', 'no_sk', 'keterangan')
    ordering = ('-tmt_jabatan',)


class PendidikanInline(admin.TabularInline):
    model   = RiwayatPendidikanTendik
    extra   = 1
    fields  = ('jenjang', 'bidang_studi', 'institusi', 'fakultas_pt', 'tahun_lulus')
    ordering = ('-jenjang',)


class DetailPPPKInline(admin.StackedInline):
    model  = DetailPPPK
    extra  = 0
    fields = ('tmt_pppk', 'tmt_berkala', 'thn_pengangkatan', 'thn_berjalan')
    verbose_name_plural = 'Detail PPPK'


class CutiInline(admin.TabularInline):
    model   = CutiTendik
    extra   = 0
    fields  = ('jenis_cuti', 'tanggal_mulai', 'tanggal_akhir', 'lama_hari_kerja', 'no_sk')
    ordering = ('-tanggal_mulai',)


class BerhentiInline(admin.StackedInline):
    model  = RiwayatBerhentiTendik
    extra  = 0
    fields = ('alasan', 'tanggal', 'no_sk', 'no_telp_keluarga', 'keterangan')
    verbose_name_plural = 'Data Berhenti / Pensiun / Meninggal'


# ─────────────────────────────────────────────────────────────────────────────
# FILTER KUSTOM
# ─────────────────────────────────────────────────────────────────────────────

class PensiunTahunIniFilter(admin.SimpleListFilter):
    title        = 'Pensiun Tahun Ini'
    parameter_name = 'pensiun_tahun_ini'

    def lookups(self, request, model_admin):
        return [('ya', f'Pensiun Tahun {date.today().year}')]

    def queryset(self, request, queryset):
        if self.value() == 'ya':
            tahun = date.today().year
            # Ambil semua, filter di Python (karena tmt_pensiun adalah @property)
            ids = [
                t.pk for t in queryset
                if t.tmt_pensiun and t.tmt_pensiun.year == tahun
            ]
            return queryset.filter(pk__in=ids)
        return queryset


class Pensiun5TahunFilter(admin.SimpleListFilter):
    title        = 'Pensiun 5 Tahun ke Depan'
    parameter_name = 'pensiun_5thn'

    def lookups(self, request, model_admin):
        return [('ya', '5 Tahun ke Depan')]

    def queryset(self, request, queryset):
        if self.value() == 'ya':
            tahun_max = date.today().year + 5
            ids = [
                t.pk for t in queryset
                if t.tmt_pensiun and t.tmt_pensiun.year <= tahun_max
            ]
            return queryset.filter(pk__in=ids)
        return queryset


# ─────────────────────────────────────────────────────────────────────────────
# TENDIK ADMIN (UTAMA)
# ─────────────────────────────────────────────────────────────────────────────

@admin.register(Tendik)
class TendikAdmin(admin.ModelAdmin):

    # ── List view ─────────────────────────────────────────────────────────────
    list_display = (
        'nip', 'nama_lengkap_display', 'status', 'nama_unit',
        'kepangkatan_display', 'jabatan_display',
        'usia_display', 'tmt_pensiun_display', 'sisa_mk_display',
    )
    list_display_links = ('nip', 'nama_lengkap_display')
    list_per_page      = 25

    # ── Search ────────────────────────────────────────────────────────────────
    search_fields = (
        'nama_terang', 'nip', 'no_ktp', 'karpeg', 'nuptk', 'nidn',
        'unit_kerja__nama', 'bagian',
    )

    # ── Filter ────────────────────────────────────────────────────────────────
    list_filter = (
        'status',
        'jenis_kelamin',
        'agama',
        'tingkat_ijazah_bkn',
        'usia_pensiun',
        'unit_kerja',
        PensiunTahunIniFilter,
        Pensiun5TahunFilter,
    )

    # ── Actions ───────────────────────────────────────────────────────────────
    actions = [
        export_tendik_excel,
        export_tendik_kepangkatan_excel,
        export_tendik_pensiun_excel,
    ]

    # ── Form layout ───────────────────────────────────────────────────────────
    fieldsets = (
        ('Identitas', {
            'fields': (
                ('nip', 'nidn', 'nuptk'),
                ('karpeg', 'nira', 'no_ktp'),
            )
        }),
        ('Nama', {
            'fields': (
                ('gelar_depan', 'nama_terang', 'gelar_belakang'),
                'nama_lengkap_readonly',
            ),
            'description': 'nama_lengkap dihitung otomatis dari gelar + nama terang'
        }),
        ('Data Pribadi', {
            'fields': (
                ('jenis_kelamin', 'agama'),
                ('tempat_lahir', 'tanggal_lahir'),
                ('no_hp', 'email'),
                'alamat',
            )
        }),
        ('Unit Kerja', {
            'fields': (
                'status',
                ('unit_kerja', 'bagian'),
            )
        }),
        ('Kepegawaian', {
            'fields': (
                ('tmt_cpns', 'tmt_pns', 'tmt_kgb'),
            )
        }),
        ('Pendidikan & Kompetensi', {
            'fields': (
                ('tingkat_ijazah_bkn', 'tingkat_ijazah_profesi', 'tingkat_ijazah_borang'),
                'bidang_keahlian',
            )
        }),
        ('Pensiun', {
            'fields': (
                'usia_pensiun',
                ('tmt_pensiun_readonly', 'tahun_pensiun_readonly', 'sisa_mk_readonly'),
            )
        }),
        ('Kalkulasi Masa Kerja (otomatis)', {
            'classes': ('collapse',),
            'fields': (
                ('mk_keseluruhan_readonly', 'mk_golongan_readonly', 'mk_jabatan_readonly'),
                'usia_saat_ini_readonly',
            )
        }),
        ('Metadata', {
            'classes': ('collapse',),
            'fields': ('created_by', 'created_at', 'updated_at')
        }),
    )

    readonly_fields = (
        'nama_lengkap_readonly',
        'tmt_pensiun_readonly',
        'tahun_pensiun_readonly',
        'sisa_mk_readonly',
        'usia_saat_ini_readonly',
        'mk_keseluruhan_readonly',
        'mk_golongan_readonly',
        'mk_jabatan_readonly',
        'created_at',
        'updated_at',
    )

    # ── Inline ────────────────────────────────────────────────────────────────
    inlines = [
        KepangkatanInline,
        JabatanFungsionalInline,
        JabatanStrukturalInline,
        TugasTambahanInline,
        PendidikanInline,
        DetailPPPKInline,
        CutiInline,
        BerhentiInline,
    ]

    # ── Readonly field methods ─────────────────────────────────────────────────
    @admin.display(description='Nama Lengkap')
    def nama_lengkap_readonly(self, obj):
        return obj.nama_lengkap

    @admin.display(description='TMT Pensiun')
    def tmt_pensiun_readonly(self, obj):
        return obj.tmt_pensiun

    @admin.display(description='Tahun Pensiun')
    def tahun_pensiun_readonly(self, obj):
        return obj.tahun_pensiun

    @admin.display(description='Sisa Masa Kerja')
    def sisa_mk_readonly(self, obj):
        return obj.sisa_masa_kerja_str

    @admin.display(description='Usia Saat Ini')
    def usia_saat_ini_readonly(self, obj):
        return f"{obj.usia_saat_ini} tahun" if obj.usia_saat_ini else '-'

    @admin.display(description='MK Keseluruhan')
    def mk_keseluruhan_readonly(self, obj):
        t, b = obj.masa_kerja_keseluruhan
        return f"{t} Thn {b} Bln"

    @admin.display(description='MK Golongan')
    def mk_golongan_readonly(self, obj):
        t, b = obj.masa_kerja_golongan
        return f"{t} Thn {b} Bln"

    @admin.display(description='MK Jabatan')
    def mk_jabatan_readonly(self, obj):
        t, b = obj.masa_kerja_jabatan
        return f"{t} Thn {b} Bln"

    # ── List display methods ───────────────────────────────────────────────────
    @admin.display(description='Nama Lengkap', ordering='nama_terang')
    def nama_lengkap_display(self, obj):
        return obj.nama_lengkap

    @admin.display(description='Pangkat / Gol')
    def kepangkatan_display(self, obj):
        kpk = obj.kepangkatan_terakhir
        if kpk:
            return format_html('<span title="{}">{}</span>', kpk.pangkat, kpk.golongan)
        return '-'

    @admin.display(description='Jabatan')
    def jabatan_display(self, obj):
        jabf = obj.jabatan_fungsional_terakhir
        if jabf:
            return jabf.nama_dan_jenjang
        jabs = obj.jabatan_struktural_aktif
        if jabs:
            return jabs.jabatan
        return '-'

    @admin.display(description='Usia')
    def usia_display(self, obj):
        u = obj.usia_saat_ini
        return f"{u} thn" if u else '-'

    @admin.display(description='TMT Pensiun')
    def tmt_pensiun_display(self, obj):
        tmt = obj.tmt_pensiun
        if not tmt:
            return '-'
        today = date.today()
        sisa_hari = (tmt - today).days
        if sisa_hari <= 0:
            return format_html('<span style="color:gray">Sudah Pensiun</span>')
        elif sisa_hari <= 365:
            return format_html('<span style="color:red;font-weight:bold">{}</span>', tmt)
        elif sisa_hari <= 365 * 2:
            return format_html('<span style="color:orange">{}</span>', tmt)
        return tmt

    @admin.display(description='Sisa MK')
    def sisa_mk_display(self, obj):
        return obj.sisa_masa_kerja_str

    # ── Optimasi query ────────────────────────────────────────────────────────
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('unit_kerja', 'created_by'
        ).prefetch_related(
            'riwayat_kepangkatan',
            'riwayat_jabatan_fungsional',
            'jabatan_struktural',
        )

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


# ─────────────────────────────────────────────────────────────────────────────
# UNIT KERJA
# ─────────────────────────────────────────────────────────────────────────────

@admin.register(UnitKerja)
class UnitKerjaAdmin(admin.ModelAdmin):
    list_display   = ('kode', 'nama', 'jumlah_tendik')
    search_fields  = ('kode', 'nama')
    ordering       = ('nama',)

    @admin.display(description='Jml Tendik')
    def jumlah_tendik(self, obj):
        return obj.tendik.count()


# ─────────────────────────────────────────────────────────────────────────────
# REFERENSI JABATAN
# ─────────────────────────────────────────────────────────────────────────────

@admin.register(ReferensiJabatan)
class ReferensiJabatanAdmin(admin.ModelAdmin):
    list_display  = ('jenis', 'nama_jabatan', 'kelas_jabatan', 'id_grade', 'nominal_gaji_display')
    list_filter   = ('jenis',)
    search_fields = ('nama_jabatan',)
    ordering      = ('jenis', 'nama_jabatan')

    @admin.display(description='Nominal Gaji', ordering='nominal_gaji')
    def nominal_gaji_display(self, obj):
        if obj.nominal_gaji:
            return f"Rp {obj.nominal_gaji:,.0f}".replace(',', '.')
        return '-'


# ─────────────────────────────────────────────────────────────────────────────
# RIWAYAT KEPANGKATAN (standalone)
# ─────────────────────────────────────────────────────────────────────────────

@admin.register(RiwayatKepangkatanTendik)
class RiwayatKepangkatanAdmin(admin.ModelAdmin):
    list_display  = ('tendik', 'pangkat', 'golongan', 'tmt', 'masa_kerja_gol')
    list_filter   = ('golongan',)
    search_fields = ('tendik__nama_terang', 'tendik__nip', 'pangkat', 'golongan')
    ordering      = ('-tmt',)
    raw_id_fields = ('tendik',)

    @admin.display(description='MK Golongan')
    def masa_kerja_gol(self, obj):
        from dateutil.relativedelta import relativedelta
        r = relativedelta(date.today(), obj.tmt)
        return f"{r.years} Thn {r.months} Bln"


# ─────────────────────────────────────────────────────────────────────────────
# JABATAN FUNGSIONAL (standalone)
# ─────────────────────────────────────────────────────────────────────────────

@admin.register(RiwayatJabatanFungsionalTendik)
class RiwayatJabatanFungsionalAdmin(admin.ModelAdmin):
    list_display  = ('tendik', 'jenis', 'nama_jabatan', 'jenjang', 'tmt', 'angka_kredit')
    list_filter   = ('jenis', 'jenjang')
    search_fields = ('tendik__nama_terang', 'tendik__nip', 'nama_jabatan')
    ordering      = ('-tmt',)
    raw_id_fields = ('tendik',)


# ─────────────────────────────────────────────────────────────────────────────
# JABATAN STRUKTURAL (standalone)
# ─────────────────────────────────────────────────────────────────────────────

@admin.register(JabatanStrukturalTendik)
class JabatanStrukturalAdmin(admin.ModelAdmin):
    list_display  = ('tendik', 'jabatan', 'eselon', 'unit_kerja', 'tmt_jabatan', 'is_aktif')
    list_filter   = ('eselon', 'is_aktif')
    search_fields = ('tendik__nama_terang', 'tendik__nip', 'jabatan', 'unit_kerja')
    ordering      = ('-tmt_jabatan',)
    raw_id_fields = ('tendik',)


# ─────────────────────────────────────────────────────────────────────────────
# CUTI (standalone)
# ─────────────────────────────────────────────────────────────────────────────

@admin.register(CutiTendik)
class CutiTendikAdmin(admin.ModelAdmin):
    list_display  = ('tendik', 'jenis_cuti', 'tanggal_mulai', 'tanggal_akhir',
                     'lama_hari_kerja', 'no_sk')
    list_filter   = ('jenis_cuti',)
    search_fields = ('tendik__nama_terang', 'tendik__nip', 'no_sk', 'alasan')
    ordering      = ('-tanggal_mulai',)
    raw_id_fields = ('tendik',)
    date_hierarchy = 'tanggal_mulai'


# ─────────────────────────────────────────────────────────────────────────────
# RIWAYAT BERHENTI (standalone)
# ─────────────────────────────────────────────────────────────────────────────

@admin.register(RiwayatBerhentiTendik)
class RiwayatBerhentiAdmin(admin.ModelAdmin):
    list_display  = ('tendik', 'alasan', 'tanggal', 'no_sk', 'no_telp_keluarga')
    list_filter   = ('alasan',)
    search_fields = ('tendik__nama_terang', 'tendik__nip', 'no_sk')
    ordering      = ('-tanggal',)
    raw_id_fields = ('tendik',)
    date_hierarchy = 'tanggal'


# ─────────────────────────────────────────────────────────────────────────────
# DETAIL PPPK (standalone)
# ─────────────────────────────────────────────────────────────────────────────

@admin.register(DetailPPPK)
class DetailPPPKAdmin(admin.ModelAdmin):
    list_display  = ('tendik', 'tmt_pppk', 'tmt_berkala',
                     'thn_pengangkatan', 'thn_berjalan', 'masa_kontrak_display')
    search_fields = ('tendik__nama_terang', 'tendik__nip')
    ordering      = ('tendik__nama_terang',)
    raw_id_fields = ('tendik',)

    @admin.display(description='MK Kontrak')
    def masa_kontrak_display(self, obj):
        t, b = obj.masa_kontrak_berjalan
        return f"{t} Thn {b} Bln"