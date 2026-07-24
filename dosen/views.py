from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import OuterRef, Subquery, Prefetch, Q, Count
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import *
from .forms import *


def is_admin(user):
    return user.is_staff or user.is_superuser


# ─── AUTH ─────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        messages.error(request, 'Username atau password salah.')
    return render(request, 'dosen/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    search = request.GET.get('q', '')
    fakultas_id = request.GET.get('fakultas', '')
    status_filter = request.GET.get('status', '')
    jabatan_filter = request.GET.get('jabatan', '')

    STATUS_TIDAK_AKTIF = ['PENSIUN', 'MENINGGAL', 'BERHENTI', 'PINDAH']
    today = timezone.localdate()

    dosen_qs = Dosen.objects.select_related('fakultas').prefetch_related(
        'riwayat_kepangkatan', 'riwayat_jabatan_fungsional',
        'tugas_tambahan', 'masa_kerja', 'riwayat_status'
    )

    if search:
        dosen_qs = dosen_qs.filter(
            Q(nama_lengkap__icontains=search) |
            Q(nuptk__icontains=search) |
            Q(nip__icontains=search) |
            Q(nama_terang__icontains=search)
        )
    if fakultas_id:
        dosen_qs = dosen_qs.filter(fakultas_id=fakultas_id)
    if status_filter:
        dosen_qs = dosen_qs.filter(status=status_filter)

    # Build dashboard rows with latest data
    rows = []
    no = 1
    for d in dosen_qs:
        status_terakhir = d.status_terakhir

        # Keluarkan dari Dashboard jika:
        # 1) sudah tercatat status tidak aktif resmi (SK sudah diproses admin), ATAU
        # 2) sudah mencapai TMT Pensiun (BUP) — meskipun SK BELUM diproses,
        #    dosen tetap otomatis pindah ke halaman Dosen Berhenti dengan status "Belum Diproses"
        if status_terakhir and status_terakhir.status in STATUS_TIDAK_AKTIF:
            continue
        if d.sudah_bup:
            continue

        kp = d.kepangkatan_terakhir
        jf = d.jabatan_fungsional_terakhir
        tt = d.tugas_tambahan_aktif
        mk = getattr(d, 'masa_kerja', None)

        if jabatan_filter and (not jf or jabatan_filter.lower() not in jf.jenjang.lower()):
            continue

        rows.append({
            'no': no,
            'dosen': d,
            'kepangkatan': kp,
            'jabatan_fungsional': jf,
            'tugas_tambahan': tt,
            'masa_kerja': mk,
        })
        no += 1

    per_page = int(request.GET.get('per_page', 10))
    if per_page > 9000:
        per_page = len(rows)
    paginator = Paginator(rows, per_page or 10)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Stats — hanya dosen aktif (bukan pensiun/meninggal/berhenti/pindah, dan belum lewat BUP)
    dosen_tidak_aktif_ids = set(
        RiwayatStatusDosen.objects.filter(
            status__in=STATUS_TIDAK_AKTIF
        ).values_list('dosen_id', flat=True)
    )
    dosen_sudah_bup_ids = set(
        Dosen.objects.filter(
            tmt_pensiun__isnull=False, tmt_pensiun__lte=today
        ).values_list('pk', flat=True)
    )
    exclude_ids = dosen_tidak_aktif_ids | dosen_sudah_bup_ids

    total = Dosen.objects.exclude(pk__in=exclude_ids).count()
    jabatan_counts = RiwayatJabatanFungsional.objects.exclude(
        dosen_id__in=exclude_ids
    ).aggregate(
        gb_count=Count('dosen', filter=Q(jenjang__icontains='Guru Besar'), distinct=True),
        lk_count=Count('dosen', filter=Q(jenjang__icontains='Lektor Kepala'), distinct=True),
        lektor_count=Count('dosen', filter=Q(jenjang__icontains='Lektor') & ~Q(jenjang__icontains='Lektor Kepala'), distinct=True),
        asisten_ahli_count=Count('dosen', filter=Q(jenjang__icontains='Asisten Ahli'), distinct=True),
        tenaga_pengajar_count=Count('dosen', filter=Q(jenjang__icontains='Tenaga Pengajar'), distinct=True)
    )

    context = {
        'page_obj': page_obj,
        'per_page': str(per_page),
        'total_dosen': total,
        'gb_count': jabatan_counts['gb_count'],
        'lk_count': jabatan_counts['lk_count'],
        'lektor_count': jabatan_counts['lektor_count'],
        'asisten_ahli_count': jabatan_counts['asisten_ahli_count'],
        'tenaga_pengajar_count': jabatan_counts['tenaga_pengajar_count'],
        'fakultas_list': Fakultas.objects.all(),
        'search': search,
        'fakultas_id': fakultas_id,
        'status_filter': status_filter,
        'jabatan_filter': jabatan_filter,
        'is_admin': is_admin(request.user),
    }
    return render(request, 'dosen/dashboard.html', context)
# ─── KENAIKAN PANGKAT ─────────────────────────────────────────────────────────

@login_required
def naik_pangkat_list(request):
    search = request.GET.get('q', '')
    fakultas_id = request.GET.get('fakultas', '')

    today = timezone.localdate()
    try:
        batas_tmt = today.replace(year=today.year - 2)
    except ValueError:
        # antisipasi 29 Feb
        batas_tmt = today.replace(month=2, day=28, year=today.year - 2)

    dosen_qs = Dosen.objects.select_related('fakultas').prefetch_related(
        'riwayat_kepangkatan', 'riwayat_jabatan_fungsional'
    )

    if search:
        dosen_qs = dosen_qs.filter(
            Q(nama_lengkap__icontains=search) |
            Q(nuptk__icontains=search) |
            Q(nip__icontains=search) |
            Q(nama_terang__icontains=search)
        )
    if fakultas_id:
        dosen_qs = dosen_qs.filter(fakultas_id=fakultas_id)

    def lama_menjabat(tmt, hari_ini):
        years = hari_ini.year - tmt.year
        months = hari_ini.month - tmt.month
        if hari_ini.day < tmt.day:
            months -= 1
        if months < 0:
            years -= 1
            months += 12
        return years, months

    def lama_menjabat(tmt, hari_ini):
        years = hari_ini.year - tmt.year
        months = hari_ini.month - tmt.month
        if hari_ini.day < tmt.day:
            months -= 1
        if months < 0:
            years -= 1
            months += 12
        return years, months

    rows = []
    no = 1
    count_2th = 0   # >= 2 tahun (mendesak)
    count_3th = 0   # >= 3 tahun (sangat mendesak)
    count_5th = 0   # >= 5 tahun (kritis)

    for d in dosen_qs:
        kp = d.kepangkatan_terakhir
        if not kp or not kp.tmt:
            continue
        if kp.tmt > batas_tmt:
            continue

        lama_tahun, lama_bulan = lama_menjabat(kp.tmt, today)

        count_2th += 1
        if lama_tahun >= 3:
            count_3th += 1
        if lama_tahun >= 5:
            count_5th += 1

        rows.append({
            'no': no,
            'dosen': d,
            'kepangkatan': kp,
            'jabatan_fungsional': d.jabatan_fungsional_terakhir,
            'lama_tahun': lama_tahun,
            'lama_bulan': lama_bulan,
        })
        no += 1

    per_page = int(request.GET.get('per_page', 25))
    if per_page > 9000:
        per_page = len(rows) or 1
    paginator = Paginator(rows, per_page or 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_eligible': len(rows),
        'count_2th': count_2th,
        'count_3th': count_3th,
        'count_5th': count_5th,
        'fakultas_list': Fakultas.objects.all(),
        'search': search,
        'fakultas_id': fakultas_id,
        'is_admin': is_admin(request.user),
    }
    return render(request, 'dosen/naik_pangkat_list.html', context)
# EXPORT DOSEN

# ─── DOSEN BERHENTI / PENSIUN (BUP) ────────────────────────────────────────────

def _tambah_tahun(tgl, tahun):
    """Tambahkan sejumlah tahun ke tanggal, aman untuk 29 Februari."""
    try:
        return tgl.replace(year=tgl.year + tahun)
    except ValueError:
        return tgl.replace(month=2, day=28, year=tgl.year + tahun)


def _hitung_bup(dosen):
    """Tentukan usia BUP: pakai field usia_pensiun jika diisi manual,
    jika tidak: Guru Besar = 70 tahun, selain itu = 65 tahun."""
    if dosen.usia_pensiun:
        return dosen.usia_pensiun
    jf = dosen.jabatan_fungsional_terakhir
    if jf and 'guru besar' in jf.jenjang.lower():
        return 70
    return 65


@login_required
def dosen_berhenti_list(request):
    search = request.GET.get('q', '')
    fakultas_id = request.GET.get('fakultas', '')

    today = timezone.localdate()

    dosen_qs = Dosen.objects.select_related('fakultas').prefetch_related(
        'riwayat_jabatan_fungsional', 'riwayat_status'
    )

    if search:
        dosen_qs = dosen_qs.filter(
            Q(nama_lengkap__icontains=search) |
            Q(nuptk__icontains=search) |
            Q(nip__icontains=search) |
            Q(nama_terang__icontains=search)
        )
    if fakultas_id:
        dosen_qs = dosen_qs.filter(fakultas_id=fakultas_id)

    rows = []
    no = 1
    count_belum_diproses = 0
    count_sudah_diproses = 0

    for d in dosen_qs:
        if not d.tanggal_lahir:
            continue

        bup = _hitung_bup(d)
        tanggal_bup = _tambah_tahun(d.tanggal_lahir, bup)

        # Belum mencapai BUP -> lewati (masih aktif, tidak masuk daftar berhenti)
        if tanggal_bup > today:
            continue

        status_terakhir = d.status_terakhir
        sudah_diproses = bool(status_terakhir and status_terakhir.status in ['PENSIUN', 'MENINGGAL', 'BERHENTI', 'PINDAH'])

        if sudah_diproses:
            count_sudah_diproses += 1
        else:
            count_belum_diproses += 1

        selisih_bulan = (today.year - tanggal_bup.year) * 12 + (today.month - tanggal_bup.month)
        if today.day < tanggal_bup.day:
            selisih_bulan -= 1
        lewat_tahun, lewat_bulan = divmod(max(selisih_bulan, 0), 12)

        rows.append({
            'no': no,
            'dosen': d,
            'bup': bup,
            'tanggal_bup': tanggal_bup,
            'lewat_tahun': lewat_tahun,
            'lewat_bulan': lewat_bulan,
            'jabatan_fungsional': d.jabatan_fungsional_terakhir,
            'status_terakhir': status_terakhir,
            'sudah_diproses': sudah_diproses,
        })
        no += 1

    # Urutkan: yang belum diproses & paling lama lewat BUP di atas
    rows.sort(key=lambda r: (r['sudah_diproses'], -(r['lewat_tahun'] * 12 + r['lewat_bulan'])))
    for i, r in enumerate(rows, start=1):
        r['no'] = i

    per_page = int(request.GET.get('per_page', 25))
    if per_page > 9000:
        per_page = len(rows) or 1
    paginator = Paginator(rows, per_page or 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_berhenti': len(rows),
        'count_belum_diproses': count_belum_diproses,
        'count_sudah_diproses': count_sudah_diproses,
        'fakultas_list': Fakultas.objects.all(),
        'search': search,
        'fakultas_id': fakultas_id,
        'is_admin': is_admin(request.user),
    }
    return render(request, 'dosen/dosen_berhenti_list.html', context)


# EXPORT DOSEN

@login_required
def export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dosen"

    # ── Styles ──────────────────────────────────────────────────────────────
    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    hdr_fill   = PatternFill("solid", start_color="1F4E79")
    sub_fill   = PatternFill("solid", start_color="374151")   # sub-header (Thn/Bln)
    sub_font   = Font(name="Arial", bold=True, color="FFFFFF", size=8)
    alt_fill   = PatternFill("solid", start_color="EBF3FB")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right      = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    thin       = Side(style="thin", color="BFBFBF")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(row, col, value, align=center):
        c = ws.cell(row=row, column=col, value=value)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = align; c.border = border
        return c

    def scell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = sub_font; c.fill = sub_fill
        c.alignment = center; c.border = border
        return c

    # ── Baris 1: Judul ──────────────────────────────────────────────────────
    TOTAL_COLS = 41
    ws.merge_cells(f"A1:{get_column_letter(TOTAL_COLS)}1")
    tc = ws["A1"]
    tc.value     = "DATA DOSEN UNIVERSITAS TADULAKO"
    tc.font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    tc.alignment = center
    ws.row_dimensions[1].height = 28
    single_cols = {
        1:  ("NO",              6),
        2:  ("NIDN",           16),
        3:  ("NIP",            20),
        4:  ("NAMA DOSEN",     32),
        5:  ("FAKULTAS",       22),
        6:  ("STATUS",         10),
        7:  ("L/P",            10),
        8:  ("PANGKAT",        20),
        9:  ("GOL.",            8),
        10: ("TMT",            13),
        11: ("JAB. FUNGSIONAL",20),
        12: ("TMT JAB.",       13),
        13: ("AK (KUM)",       10),
        24: ("JURUSAN/BAGIAN", 22),
        25: ("PRODI",          22),
        26: ("NO. REG SERDOS", 18),
        27: ("KARPEG",         14),
        28: ("TMT CPNS",       13),
        29: ("TMT PNS",        13),
        30: ("GELAR DEPAN",    14),
        31: ("GELAR BELAKANG", 16),
        32: ("AGAMA",          12),
        33: ("KEPAKARAN",      22),
        34: ("TUGAS TAMBAHAN", 24),
        35: ("IJ. BKN",        10),
        36: ("IJ. BORANG",     11),
        37: ("TMP LAHIR",      16),
        38: ("TGL LAHIR",      13),
        39: ("USIA",            7),
        40: ("PENSIUN",         9),
        41: ("TMT PENSIUN",    14),
    }

    # Kolom dengan sub-header Thn/Bln (merge baris 2, sub di baris 3)
    grouped_cols = {
        14: "MASA KERJA CPNS",
        16: "MASA KERJA GOL",
        18: "MASA KERJA JAB",
        20: "MK KESELURUHAN",
        22: "MK PENSIUN",
    }

    # Render single cols (merge row 2-3)
    for col, (label, width) in single_cols.items():
        ws.merge_cells(start_row=2, start_column=col,
                       end_row=3,   end_column=col)
        hcell(2, col, label)
        ws.column_dimensions[get_column_letter(col)].width = width

    # Render grouped cols (merge 2 kolom di row 2, lalu Thn/Bln di row 3)
    for start_col, label in grouped_cols.items():
        ws.merge_cells(start_row=2, start_column=start_col,
                       end_row=2,   end_column=start_col + 1)
        hcell(2, start_col, label)
        scell(3, start_col,     "Thn")
        scell(3, start_col + 1, "Bln")
        ws.column_dimensions[get_column_letter(start_col)].width     = 7
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 7

    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"

    # ── Data rows (mulai baris 4) ────────────────────────────────────────────
    qs = (
        Dosen.objects
        .select_related("fakultas")
        .prefetch_related(
            "riwayat_kepangkatan",
            "riwayat_jabatan_fungsional",
            "tugas_tambahan",
            "masa_kerja",
        )
        .order_by("nama_lengkap")
    )

    def fmt_date(d):
        return d.strftime("%d-%m-%Y") if d else "-"

    def val_or_dash(v):
        return v if v else "-"

    for row_num, dosen in enumerate(qs, start=1):
        row_idx = row_num + 3
        fill    = alt_fill if row_num % 2 == 0 else None

        kpk  = dosen.kepangkatan_terakhir          # RiwayatKepangkatan | None
        jab  = dosen.jabatan_fungsional_terakhir   # RiwayatJabatanFungsional | None
        tt   = dosen.tugas_tambahan_aktif          # TugasTambahan | None
        mk   = getattr(dosen, "masa_kerja", None)  # MasaKerja | None

        row_data = [
            # col 1-13
            row_num,
            val_or_dash(dosen.nidn),
            val_or_dash(dosen.nip),
            dosen.nama_lengkap,
            dosen.fakultas.nama if dosen.fakultas else "-",
            dosen.get_status_display(),
            dosen.get_jenis_kelamin_display(),
            kpk.pangkat if kpk else "-",
            kpk.golongan if kpk else "-",
            fmt_date(kpk.tmt) if kpk else "-",
            jab.jenjang if jab else "-",
            fmt_date(jab.tmt) if jab else "-",
            float(jab.angka_kredit) if (jab and jab.angka_kredit) else "-",
            # col 14-15 MK CPNS
            mk.cpns_tahun if mk else "-",
            mk.cpns_bulan if mk else "-",
            # col 16-17 MK GOL
            mk.golongan_tahun if mk else "-",
            mk.golongan_bulan if mk else "-",
            # col 18-19 MK JAB
            mk.jabatan_tahun if mk else "-",
            mk.jabatan_bulan if mk else "-",
            # col 20-21 MK KESELURUHAN
            mk.keseluruhan_tahun if mk else "-",
            mk.keseluruhan_bulan if mk else "-",
            # col 22-23 MK PENSIUN
            mk.pensiun_tahun if mk else "-",
            mk.pensiun_bulan if mk else "-",
            # col 24-41
            val_or_dash(dosen.jurusan_bagian),
            val_or_dash(dosen.program_studi_nama),
            val_or_dash(dosen.no_reg_serdos),
            val_or_dash(dosen.karpeg),
            fmt_date(dosen.tmt_cpns),
            fmt_date(dosen.tmt_pns),
            val_or_dash(dosen.gelar_depan),
            val_or_dash(dosen.gelar_belakang),
            val_or_dash(dosen.agama),
            val_or_dash(dosen.ranting_ilmu),
            tt.jabatan if tt else "-",
            val_or_dash(dosen.tingkat_ijazah_bkn),
            val_or_dash(dosen.tingkat_ijazah_borang),
            val_or_dash(dosen.tempat_lahir),
            fmt_date(dosen.tanggal_lahir),
            dosen.usia_saat_ini or "-",
            dosen.tahun_pensiun or "-",
            fmt_date(dosen.tmt_pensiun),
        ]

        # Alignment per kolom
        aligns = [center] + [left] * 40   # default semua left kecuali NO
        aligns[0]  = center   # NO
        aligns[6]  = center   # L/P
        aligns[8]  = center   # GOL
        aligns[12] = right    # AK (KUM)
        aligns[13] = center   # MK cols (14-23)
        aligns[14] = center
        aligns[15] = center
        aligns[16] = center
        aligns[17] = center
        aligns[18] = center
        aligns[19] = center
        aligns[20] = center
        aligns[21] = center
        aligns[22] = center
        aligns[38] = center   # USIA
        aligns[39] = center   # PENSIUN
        aligns[40] = center   # TMT PENSIUN

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = aligns[col_idx - 1]
            cell.border    = border
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 25

    # ── HTTP Response ────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Data_Dosen_UNTAD.xlsx"'
    return response
# ─── DOSEN DETAIL ─────────────────────────────────────────────────────────────

@login_required
def dosen_detail(request, pk):
    dosen = get_object_or_404(Dosen.objects.select_related('fakultas').prefetch_related(
        'riwayat_kepangkatan', 'riwayat_jabatan_fungsional',
        'riwayat_pendidikan', 'tugas_tambahan', 'masa_kerja'
    ), pk=pk)
    context = {
        'dosen': dosen,
        'is_admin': is_admin(request.user),
    }
    return render(request, 'dosen/dosen_detail.html', context)


# ─── DOSEN CRUD (ADMIN ONLY) ──────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def dosen_create(request):
    if request.method == 'POST':
        form = DosenForm(request.POST)
        if form.is_valid():
            dosen = form.save(commit=False)
            dosen.created_by = request.user
            dosen.save()
            messages.success(request, f'Data dosen {dosen.nama_lengkap} berhasil disimpan.')
            return redirect('dosen_detail', pk=dosen.pk)
    else:
        form = DosenForm()
    return render(request, 'dosen/dosen_form.html', {'form': form, 'title': 'Tambah Dosen Baru'})


@login_required
@user_passes_test(is_admin)
def dosen_edit(request, pk):
    dosen = get_object_or_404(Dosen, pk=pk)
    if request.method == 'POST':
        form = DosenForm(request.POST, instance=dosen)
        if form.is_valid():
            form.save()
            messages.success(request, f'Data dosen {dosen.nama_lengkap} berhasil diperbarui.')
            return redirect('dosen_detail', pk=dosen.pk)
    else:
        form = DosenForm(instance=dosen)
    return render(request, 'dosen/dosen_form.html', {
        'form': form, 'dosen': dosen, 'title': f'Edit: {dosen.nama_lengkap}'
    })


@login_required
@user_passes_test(is_admin)
def dosen_delete(request, pk):
    dosen = get_object_or_404(Dosen, pk=pk)
    if request.method == 'POST':
        nama = dosen.nama_lengkap
        dosen.delete()
        messages.success(request, f'Data dosen {nama} berhasil dihapus.')
        return redirect('dashboard')
    return render(request, 'dosen/confirm_delete.html', {'obj': dosen, 'title': 'Hapus Dosen'})


# ─── RIWAYAT KEPANGKATAN ──────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def kepangkatan_add(request, dosen_pk):
    dosen = get_object_or_404(Dosen, pk=dosen_pk)
    if request.method == 'POST':
        form = RiwayatKepangkatanForm(request.POST)
        if form.is_valid():
            riwayat = form.save(commit=False)
            riwayat.dosen = dosen
            riwayat.save()
            messages.success(request, 'Riwayat kepangkatan berhasil ditambahkan.')
            return redirect('dosen_detail', pk=dosen_pk)
    else:
        form = RiwayatKepangkatanForm()
    return render(request, 'dosen/riwayat_form.html', {
        'form': form, 'dosen': dosen, 'title': 'Tambah Riwayat Kepangkatan',
    })


@login_required
@user_passes_test(is_admin)
def kepangkatan_edit(request, pk):
    riwayat = get_object_or_404(RiwayatKepangkatan, pk=pk)
    if request.method == 'POST':
        form = RiwayatKepangkatanForm(request.POST, instance=riwayat)
        if form.is_valid():
            form.save()
            messages.success(request, 'Riwayat kepangkatan berhasil diperbarui.')
            return redirect('dosen_detail', pk=riwayat.dosen_id)
    else:
        form = RiwayatKepangkatanForm(instance=riwayat)
    return render(request, 'dosen/riwayat_form.html', {
        'form': form, 'dosen': riwayat.dosen, 'title': 'Edit Riwayat Kepangkatan'
    })


@login_required
@user_passes_test(is_admin)
def kepangkatan_delete(request, pk):
    riwayat = get_object_or_404(RiwayatKepangkatan, pk=pk)
    dosen_pk = riwayat.dosen_id
    if request.method == 'POST':
        riwayat.delete()
        messages.success(request, 'Riwayat kepangkatan dihapus.')
        return redirect('dosen_detail', pk=dosen_pk)
    return render(request, 'dosen/confirm_delete.html', {'obj': riwayat, 'title': 'Hapus Kepangkatan'})


# ─── RIWAYAT JABATAN FUNGSIONAL ───────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def jabatan_add(request, dosen_pk):
    dosen = get_object_or_404(Dosen, pk=dosen_pk)
    if request.method == 'POST':
        form = RiwayatJabatanFungsionalForm(request.POST)
        if form.is_valid():
            riwayat = form.save(commit=False)
            riwayat.dosen = dosen
            riwayat.save()
            messages.success(request, 'Riwayat jabatan fungsional berhasil ditambahkan.')
            return redirect('dosen_detail', pk=dosen_pk)
    else:
        form = RiwayatJabatanFungsionalForm()
    return render(request, 'dosen/riwayat_form.html', {
        'form': form, 'dosen': dosen, 'title': 'Tambah Riwayat Jabatan Fungsional'
    })


@login_required
@user_passes_test(is_admin)
def jabatan_edit(request, pk):
    riwayat = get_object_or_404(RiwayatJabatanFungsional, pk=pk)
    if request.method == 'POST':
        form = RiwayatJabatanFungsionalForm(request.POST, instance=riwayat)
        if form.is_valid():
            form.save()
            messages.success(request, 'Riwayat jabatan fungsional berhasil diperbarui.')
            return redirect('dosen_detail', pk=riwayat.dosen_id)
    else:
        form = RiwayatJabatanFungsionalForm(instance=riwayat)
    return render(request, 'dosen/riwayat_form.html', {
        'form': form, 'dosen': riwayat.dosen, 'title': 'Edit Jabatan Fungsional'
    })


@login_required
@user_passes_test(is_admin)
def jabatan_delete(request, pk):
    riwayat = get_object_or_404(RiwayatJabatanFungsional, pk=pk)
    dosen_pk = riwayat.dosen_id
    if request.method == 'POST':
        riwayat.delete()
        messages.success(request, 'Riwayat jabatan fungsional dihapus.')
        return redirect('dosen_detail', pk=dosen_pk)
    return render(request, 'dosen/confirm_delete.html', {'obj': riwayat, 'title': 'Hapus Jabatan Fungsional'})


# ─── RIWAYAT PENDIDIKAN ───────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def pendidikan_add(request, dosen_pk):
    dosen = get_object_or_404(Dosen, pk=dosen_pk)
    if request.method == 'POST':
        form = RiwayatPendidikanForm(request.POST)
        if form.is_valid():
            pend = form.save(commit=False)
            pend.dosen = dosen
            pend.save()
            messages.success(request, 'Riwayat pendidikan berhasil ditambahkan.')
            return redirect('dosen_detail', pk=dosen_pk)
    else:
        form = RiwayatPendidikanForm()
    return render(request, 'dosen/riwayat_form.html', {
        'form': form, 'dosen': dosen, 'title': 'Tambah Riwayat Pendidikan'
    })


@login_required
@user_passes_test(is_admin)
def pendidikan_delete(request, pk):
    pend = get_object_or_404(RiwayatPendidikan, pk=pk)
    dosen_pk = pend.dosen_id
    if request.method == 'POST':
        pend.delete()
        messages.success(request, 'Riwayat pendidikan dihapus.')
        return redirect('dosen_detail', pk=dosen_pk)
    return render(request, 'dosen/confirm_delete.html', {'obj': pend, 'title': 'Hapus Riwayat Pendidikan'})


# ─── TUGAS TAMBAHAN ───────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def tugas_tambahan_add(request, dosen_pk):
    dosen = get_object_or_404(Dosen, pk=dosen_pk)
    if request.method == 'POST':
        form = TugasTambahanForm(request.POST)
        if form.is_valid():
            tt = form.save(commit=False)
            tt.dosen = dosen
            tt.save()
            messages.success(request, 'Tugas tambahan berhasil ditambahkan.')
            return redirect('dosen_detail', pk=dosen_pk)
    else:
        form = TugasTambahanForm()
    return render(request, 'dosen/riwayat_form.html', {
        'form': form, 'dosen': dosen, 'title': 'Tambah Tugas Tambahan'
    })


@login_required
@user_passes_test(is_admin)
def tugas_tambahan_edit(request, pk):
    tt = get_object_or_404(TugasTambahan, pk=pk)
    if request.method == 'POST':
        form = TugasTambahanForm(request.POST, instance=tt)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tugas tambahan berhasil diperbarui.')
            return redirect('dosen_detail', pk=tt.dosen_id)
    else:
        form = TugasTambahanForm(instance=tt)
    return render(request, 'dosen/riwayat_form.html', {
        'form': form, 'dosen': tt.dosen, 'title': 'Edit Tugas Tambahan'
    })


@login_required
@user_passes_test(is_admin)
def tugas_tambahan_delete(request, pk):
    tt = get_object_or_404(TugasTambahan, pk=pk)
    dosen_pk = tt.dosen_id
    if request.method == 'POST':
        tt.delete()
        messages.success(request, 'Tugas tambahan dihapus.')
        return redirect('dosen_detail', pk=dosen_pk)
    return render(request, 'dosen/confirm_delete.html', {'obj': tt, 'title': 'Hapus Tugas Tambahan'})


# ─── MASA KERJA ───────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def masa_kerja_edit(request, dosen_pk):
    dosen = get_object_or_404(Dosen, pk=dosen_pk)
    mk, _ = MasaKerja.objects.get_or_create(dosen=dosen)
    if request.method == 'POST':
        form = MasaKerjaForm(request.POST, instance=mk)
        if form.is_valid():
            form.save()
            messages.success(request, 'Masa kerja berhasil diperbarui.')
            return redirect('dosen_detail', pk=dosen_pk)
    else:
        form = MasaKerjaForm(instance=mk)
    return render(request, 'dosen/riwayat_form.html', {
        'form': form, 'dosen': dosen, 'title': 'Edit Masa Kerja'
    })


# ─── FAKULTAS MANAGEMENT ──────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def fakultas_list(request):
    fakultas_qs = Fakultas.objects.annotate(jumlah_dosen=Count('dosen')).order_by('nama')
    return render(request, 'dosen/fakultas_list.html', {'fakultas_list': fakultas_qs})


@login_required
@user_passes_test(is_admin)
def fakultas_create(request):
    if request.method == 'POST':
        form = FakultasForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fakultas berhasil ditambahkan.')
            return redirect('fakultas_list')
    else:
        form = FakultasForm()
    return render(request, 'dosen/riwayat_form.html', {'form': form, 'title': 'Tambah Fakultas'})


@login_required
@user_passes_test(is_admin)
def fakultas_edit(request, pk):
    fak = get_object_or_404(Fakultas, pk=pk)
    if request.method == 'POST':
        form = FakultasForm(request.POST, instance=fak)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fakultas berhasil diperbarui.')
            return redirect('fakultas_list')
    else:
        form = FakultasForm(instance=fak)
    return render(request, 'dosen/riwayat_form.html', {'form': form, 'title': f'Edit Fakultas: {fak.nama}'})


@login_required
def status_dosen_add(request, dosen_id):
    dosen = get_object_or_404(Dosen, pk=dosen_id)

    if request.method == "POST":
        form = RiwayatStatusDosenForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.dosen = dosen
            obj.save()

            messages.success(request, "Riwayat status berhasil ditambahkan.")
            return redirect('dosen_detail', pk=dosen.pk)
    else:
        form = RiwayatStatusDosenForm()

    return render(request, 'dosen/riwayat_form.html', {
        'form': form,
        'dosen': dosen,
        'judul': 'Tambah Riwayat Status'
    })


@login_required
def status_dosen_edit(request, pk):
    obj = get_object_or_404(RiwayatStatusDosen, pk=pk)

    if request.method == "POST":
        form = RiwayatStatusDosenForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()

            messages.success(request, "Riwayat status berhasil diperbarui.")
            return redirect('dosen_detail', pk=obj.dosen.pk)
    else:
        form = RiwayatStatusDosenForm(instance=obj)

    return render(request, 'dosen/riwayat_form.html', {
        'form': form,
        'dosen': obj.dosen,
        'judul': 'Edit Riwayat Status'
    })


@login_required
def status_dosen_delete(request, pk):
    obj = get_object_or_404(RiwayatStatusDosen, pk=pk)

    if request.method == 'POST':
        dosen_pk = obj.dosen.pk
        obj.delete()

        messages.success(request, "Riwayat status berhasil dihapus.")
        return redirect('dosen_detail', pk=dosen_pk)

    return render(request, 'dosen/confirm_delete.html', {
        'obj': obj,
        'judul': 'Hapus Riwayat Status'
    })

@login_required
@user_passes_test(is_admin)
def keluarga_add(request, dosen_id):
    dosen = get_object_or_404(Dosen, pk=dosen_id)

    if request.method == "POST":
        form = KeluargaDosenForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.dosen = dosen
            obj.save()
            messages.success(request, "Data keluarga berhasil ditambahkan.")
            return redirect('dosen_detail', pk=dosen.pk)
    else:
        form = KeluargaDosenForm()

    return render(request, 'dosen/riwayat_form.html', {
        'form': form,
        'dosen': dosen,
        'title': 'Tambah Data Keluarga'
    })


@login_required
@user_passes_test(is_admin)
def keluarga_edit(request, pk):
    obj = get_object_or_404(KeluargaDosen, pk=pk)

    if request.method == "POST":
        form = KeluargaDosenForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Data keluarga berhasil diperbarui.")
            return redirect('dosen_detail', pk=obj.dosen.pk)
    else:
        form = KeluargaDosenForm(instance=obj)

    return render(request, 'dosen/riwayat_form.html', {
        'form': form,
        'dosen': obj.dosen,
        'title': 'Edit Data Keluarga'
    })


@login_required
def keluarga_delete(request, pk):
    obj = get_object_or_404(KeluargaDosen, pk=pk)

    if request.method == 'POST':
        dosen_pk = obj.dosen.pk
        obj.delete()
        messages.success(request, "Data keluarga berhasil dihapus.")
        return redirect('dosen_detail', pk=dosen_pk)

    return render(request, 'dosen/confirm_delete.html', {
        'obj': obj,
        'title': 'Hapus Data Keluarga'
    })